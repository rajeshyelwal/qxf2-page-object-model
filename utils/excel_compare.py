"""
Utility script to compare two excel file using openxl module
"""

import openpyxl
import os

class Excel_Compare():    
    def is_equal(self,xl_actual,xl_expected):
        "Method to compare the Actual and Expected xl file"
        result_flag = True
        if not os.path.exists(xl_actual):
            result_flag = False
            print 'Could not locate the excel file: %s'%xl_actual

        if not os.path.exists(xl_expected):
            result_flag = False
            print 'Could not locate the excel file %s'%xl_expected

        if os.path.exists(xl_actual) and os.path.exists(xl_expected):
            #Open the xl file and put the content to list
            actual_xlfile = openpyxl.load_workbook(xl_actual)
            xl_sheet = actual_xlfile.active
            actual_file = []
            for row in xl_sheet.iter_rows(min_row=1, max_col=xl_sheet.max_column, max_row=xl_sheet.max_row):
                for cell in row:
                    actual_file.append(cell.value)

            exp_xlfile = openpyxl.load_workbook(xl_expected)
            xl_sheet = exp_xlfile.active
            exp_file = []
            for row in xl_sheet.iter_rows(min_row=1, max_col=xl_sheet.max_column, max_row=xl_sheet.max_row):
                for cell in row:
                    exp_file.append(cell.value)
            
            #If there is row and column mismatch result_flag = False
            if (len(actual_file)!= len(exp_file)):
                result_flag = False
                print "Mismatch in number of rows or columns. The actual row or column count didn't match with expected row or column count"
            else:
                for actual_row, actual_col in zip(actual_file,exp_file):
                    if actual_row == actual_col:
                        pass
                    else:
                        print "Mismatch between actual and expected file at position: ",actual_file.index(actual_row)
                        print "Data present only in Actual file: %s"%actual_row
                        print "Data present only in Expected file: %s"%actual_col
                        result_flag =  False

            return result_flag
        
#---USAGE EXAMPLES
if __name__=='__main__':
    print "Start of %s"%__file__
     
    file1 = 'c:/Indira/sample/excel_file1.xlsx'
    file2 = 'c:/Indira/sample/excel_file2.xlsx'
    #Initialize the csv object
    excel_obj = Excel_Compare()

    #Sample code to compare excel files
    if excel_obj.is_equal(file1,file2) is True:
        print "Data in both the excel files matched\n"
    else:
        print "Data mismatch in both the excel files"    